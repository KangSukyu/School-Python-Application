import openpyxl as px

wb = px.load_workbook("xlsx07\shift_202210.xlsx")
ws = wb.active

cnt=0

for row in ws.iter_cols(min_row=2, min_col=3):
    for cell in row:
        if cell.value == "〇":
            cnt+=1
    print(cnt*4)



for row in ws.iter_rows(max_row=1, min_col=3):
     for cell in row:
        print(f"{cell.value} {cnt*4}")

