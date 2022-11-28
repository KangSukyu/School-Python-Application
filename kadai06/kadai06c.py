import openpyxl as px
wb = px.load_workbook("xlsx\seiseki.xlsx")
ws = wb.active

Sheet_Max_Clm = ws.max_column
for i in range(1,Sheet_Max_Clm + 5):
    data = ws.cell(row=i+1, column=3).value
    fix = data+3
    if(fix >=100):
        fix = 100
    ws.cell(row=i+1, column=3).value = fix

wb.save("xlsx\seiseki_fix.xlsx")
wb.close()