import openpyxl as px
wb = px.load_workbook("xlsx\seiseki_fix.xlsx")
ws = wb.active

ws["G1"] = "合計点"
ws["H1"] = "平均点"

total = 0
# Sheet_Max_Row= ws.max_row
for i in range(1,6):
    data = ws.cell(row=2, column=i+1).value
    total = total + data

ws["G2"] = total

for i in range(1,10):
    ws.cell(row=i+1, column=8).value = total


wb.save("xlsx\seiseki_fix.xlsx")
wb.close()