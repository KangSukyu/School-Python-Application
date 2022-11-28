import openpyxl as px

wb = px.load_workbook("xlsx07\sample07_202211_customer.xlsx")
wb_item = px.load_workbook("xlsx07\sample07_price.xlsx")
ws_item = wb_item.active
customers = wb.sheetnames
for name in customers:
    ws = wb[name]
    items_dict = {}
    for i in range(1, ws.max_row+1):
        goods = ws.cell(row=i, column=2).value
        amount = ws.cell(row=i, column=3).value
        if goods in items_dict:
            items_dict[goods] += amount
        else:
            items_dict[goods] = amount
    items = []
    for item in items_dict:
        for i in range(1, ws_item.max_row+1):
            if item == ws_item.cell(row=i, column=1).value:
                price = ws_item.cell(row=i, column=2).value
                break
        items.append([item,items_dict[item], price])
    print(name)
    print(items)
wb.close()
wb_item.close()