import openpyxl as px

wb = px.load_workbook("xlsx07\sample07.xlsx")
ws = wb["注文データ"]
cells_list = []
for row in ws.iter_rows(min_row=2):
    cells = []
    for cell in row:
        cells.append(cell.value)
    cells_list.append(cells)

wb_new = px.Workbook()
ws_new = wb_new.active
ws_new.title = "注文データピックアップ"
for cnt_r, row in enumerate(cells_list, start=1):
    for cnt_c, value in enumerate(row, start=1):
        ws_new.cell(row=cnt_r, column=cnt_c).value = value

    ws_new.cell(row=cnt_r, column=4).number_format = "yyyy/mm/dd"
wb_new.save("xlsx07\sample07_pickup.xlsx")
wb_new.close()
