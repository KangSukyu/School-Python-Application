import openpyxl as px
from datetime import datetime

wb = px.load_workbook("xlsx07\sample07.xlsx")
ws = wb["注文データ"]
cells_list = []
start_day = datetime(year=2022, month=11, day=1)
end_day = datetime(year=2022, month=11, day=30)
for row in ws.iter_rows(min_row=2):
    if start_day <= row[3].value <= end_day:
        cells = []
        for cell in row:
            cells.append(cell.value)
        cells_list.append(cells)

wb_new = px.Workbook()
ws_new = wb_new.active
ws_new.title = "注文データ_202211"
for cnt_r, row in enumerate(cells_list, start=1):
    for cnt_c, value in enumerate(row, start=1):
        ws_new.cell(row=cnt_r, column=cnt_c).value = value
        
    ws_new.cell(row=cnt_r, column=4).number_format = "yyyy/mm/dd"
wb_new.save("xlsx07\sample07_202211.xlsx")
wb_new.close()

