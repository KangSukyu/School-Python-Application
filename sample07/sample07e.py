import openpyxl as px
from datetime import datetime

wb = px.load_workbook("xlsx07\sample07.xlsx")
ws = wb["注文データ"]
cells_list = []
start_day = datetime(year=2022, month=11, day=1)
end_day = datetime(year=2022, month=11, day=30)
for row in ws.iter_rows(min_row=2):
    if start_day <= row[3].value <= end_day:
        cells = []
        for cell in row:
            cells.append(cell.value)
        cells_list.append(cells)

wb_new = px.Workbook()
ws_new = wb_new.active

for row in cells_list:
    name = row[0]
    if name not in wb_new.sheetnames:
        wb_new.create_sheet(name)
    ws_target = wb_new[name]
    values = []
    for cell in row:
        values.append(cell)
    ws_target.append(values)

    max_row = ws_target.max_row
    ws_target.cell(row=max_row, column=4).number_format = "yyyy/mm/dd"
wb_new.remove(wb_new["Sheet"])
wb_new.save("xlsx07/sample07_202211_customer.xlsx")
wb_new.close()