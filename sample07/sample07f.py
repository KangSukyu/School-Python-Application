import openpyxl as px
from datetime import datetime

name = "A社"
items = [
    ["りんご", 50, 100],
    ["バナナ", 30, 200],
    ["メロン", 10, 1000]
]
wb = px.load_workbook("xlsx07/請求書_テンプレート.xlsx")
ws = wb.active

ws["I2"] = datetime.today()
ws["I2"].number_format = "yyyy年 mm月 dd日"
ws["B5"] = name
total = 0
for i, item in enumerate(items):
    goods, amount, price = item
    value = amount * price
    tax = value * 0.1
    subtotal = value + tax
    total += subtotal
    start_row = 19 + i
    ws.cell(row=start_row, column=3).value = goods
    ws.cell(row=start_row, column=7).value = amount
    ws.cell(row=start_row, column=8).value = value
    ws.cell(row=start_row, column=9).value = tax
    ws.cell(row=start_row, column=10).value = subtotal
ws["J29"] = total
ws["E15"] = total
wb.save("xlsx07/請求書_練習.xlsx")
wb.close()