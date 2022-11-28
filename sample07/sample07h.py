import openpyxl as px
from datetime import datetime

wb_data = px.load_workbook("xlsx07\sample07_202211_customer.xlsx")
wb_item = px.load_workbook("xlsx07\sample07_price.xlsx")
ws_item = wb_item.active
customers = wb_data.sheetnames
today = datetime.today()
year = datetime.today().year
month = datetime.today().month
for name in customers:
    ws_data = wb_data[name]
    items_dict = {}
    for i in range(1, ws_data.max_row+1):
        goods = ws_data.cell(row=i, column=2).value
        amount = ws_data.cell(row=i, column=3).value
        if goods in items_dict:
            items_dict[goods] += amount
        else:
            items_dict[goods] = amount
    items = []
    for item in items_dict:
        for i in range(1, ws_item.max_row+1):
            if item == ws_item.cell(row=i, column=1).value:
                price = ws_item.cell(row=i, column=2).value
                break
        items.append([item, items_dict[item], price])
    wb_temp = px.load_workbook("xlsx07\請求書_テンプレート.xlsx")
    ws_temp = wb_temp.active
    ws_temp["I2"].number_format = "yyyy年 mm月 dd日"
    ws_temp["B5"] = name
    total = 0
    for i, item in enumerate(items):
        goods, amount, price = item
        value = amount * price
        tax = value * 0.1
        subtotal = value + tax
        total += subtotal
        start_row = 19 + i
        ws_temp.cell(row=start_row, column=3).value = goods
        ws_temp.cell(row=start_row, column=7).value = amount
        ws_temp.cell(row=start_row, column=8).value = value
        ws_temp.cell(row=start_row, column=9).value = tax
        ws_temp.cell(row=start_row, column=10).value = subtotal
    ws_temp["J29"] = total
    ws_temp["E15"] = total
    wb_temp.save(f"xlsx07/請求書_{name}_{year}年{month}月.xlsx")
    wb_temp.close()
wb_data.close()
wb_item.close()