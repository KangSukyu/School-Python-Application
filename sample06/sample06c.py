from fileinput import close
import openpyxl as px
wb = px.load_workbook("xlsx/sample06_new.xlsx")

wb.create_sheet("追加シート_First",0)
wb.create_sheet("追加シート_Last")
wb.save("xlsx/sample06_new.xlsx")
close()