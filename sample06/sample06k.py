import openpyxl as px
import random
wb = px.Workbook()
ws = wb.active
ws.title = "乱数"

for i in range(1,11):
    num1 = random.randint(1,9)
    num2 = random.randint(1,9)
    ws.cell(row=i, column=1).value = num1
    ws.cell(row=i, column=2).value = num2
    ws.cell(row=i, column=3).value = num1 * num2

wb.save("xlsx/sample06_random.xlsx")
wb.close()