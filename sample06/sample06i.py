import openpyxl as px
wb = px.load_workbook("xlsx/sample06_new.xlsx")
ws = wb["夜のあいさつ"]

cell1 = ws["A2"]
cell2 = ws.cell(row=3, column=1)
print(cell1.value)
print(cell2.value)

ws["A2"] = "まだ眠くないです"
cell1 = ws["A2"]
print(cell1.value)

cell2.value = "まだまだ頑張ります"
print(cell2.value)

wb.save("xlsx/sample06_new.xlsx")
wb.close()