import openpyxl as px
wb = px.load_workbook("xlsx/sample06_new.xlsx")
ws = wb["サンプルデータ"]
max_row = ws.max_row
max_column = ws.max_column
print(ws.cell(row=max_row, column=max_column).value)