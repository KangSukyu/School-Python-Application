import openpyxl as px
wb = px.load_workbook("xlsx/sample06_new.xlsx")
ws = wb["サンプルデータ"]

cell1 = ws["B2"]
print(cell1)
print(cell1.value)

cell2 = ws.cell(row=2, column=2)
print(cell2)
print(cell2.value)
