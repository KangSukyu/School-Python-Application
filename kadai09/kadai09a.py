from re import X
from xml.etree.ElementTree import tostring
from selenium import webdriver
from selenium.webdriver.chrome import service as fs
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl as px

wb = px.load_workbook("xlsx09\出張予定.xlsx")
ws = wb['出張予定']
sheet = wb.get_sheet_by_name('出張予定')

cells_list = []
for row in ws.iter_rows():
    cells = []
    for cell in row:
        cells.append(cell.value)
    cells_list.append(cells)

list = []
for i in range(3):
    DRIVER_PASS = "web/chromedriver.exe"
    TARGET_URL = "https://transit.yahoo.co.jp/"
    DEPARTURE = "新宿駅"
    ARRIVAL = sheet.cell(row=i+3, column=2).value

    chrome_service = fs.Service(DRIVER_PASS)
    chrome_optiton = webdriver.ChromeOptions()
    chrome_optiton.add_experimental_option("excludeSwitches", ["enable-logging"])
    chrome_driver = webdriver.Chrome(service=chrome_service, options=chrome_optiton)
    chrome_driver.get(TARGET_URL)
    sleep(2)

    chrome_driver.find_element(By.NAME,"from").send_keys(DEPARTURE)
    chrome_driver.find_element(By.NAME,"to").send_keys(ARRIVAL)

    year = chrome_driver.find_element(By.ID,"y")
    year_select = Select(year)
    y = sheet.cell(row=i+3, column=1).value.strftime("%Y")
    year_select.select_by_value(y)

    month = chrome_driver.find_element(By.ID,"m")
    month_select = Select(month)
    m = sheet.cell(row=i+3, column=1).value.strftime("%m")
    month_select.select_by_value(m)

    day = chrome_driver.find_element(By.ID,"d")
    day_select = Select(day)
    d = sheet.cell(row=i+3, column=1).value.strftime("%d")
    day_select.select_by_value(d)

    hour = chrome_driver.find_element(By.ID,"hh")
    hour_select = Select(hour)
    h = sheet.cell(row=i+3, column=3).value.strftime("%H")
    hour_select.select_by_value(h)

    minute = chrome_driver.find_element(By.ID,"mm")
    minute_select = Select(minute)
    m = sheet.cell(row=i+3, column=3).value.strftime("%M")
    minute_select.select_by_value(m)

    chrome_driver.find_element(By.ID,"tsArr").click()
    chrome_driver.find_element(By.ID,"air").click()
    sleep(2)

    chrome_driver.find_element(By.ID,"searchModuleSubmit").click()
    sleep(2)

    elm = chrome_driver.find_element(By.XPATH,"//*[@id='route01']/div/div[1]/ul[1]/li")
    list.append(elm.text)

wb_new = px.Workbook()
ws_new = wb_new.active
ws_new.title = "出張予定(出発時刻追記)"
for cnt_r, row in enumerate(cells_list, start=1):
    for cnt_c, value in enumerate(row, start=1):
        ws_new.cell(row=cnt_r, column=cnt_c).value = value
    ws_new.cell(row=cnt_r, column=1).number_format = "yyyy/mm/dd"

ws_new.cell(row=2, column=4).value = '出発時刻'

for i in range(3):
    ws_new.cell(row=i+3, column=4).value = list[i]

wb_new.save("xlsx09\出張予定(出発時刻追記).xlsx")
wb_new.close()


