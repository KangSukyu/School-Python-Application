from selenium import webdriver
from selenium.webdriver.chrome import service as fs
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl as px

wb = px.load_workbook("xlsx08\search_words.xlsx")
ws = wb.active

DRIVER_PASS = "web/chromedriver.exe"
TARGET_URL = "https://www.google.co.jp/"
SEARCH_WORD = ws["A1"].value

chrome_service = fs.Service(DRIVER_PASS)
chrome_optiton = webdriver.ChromeOptions()
chrome_optiton.add_experimental_option("excludeSwitches",["enable-logging"])
chrome_driver = webdriver.Chrome(service=chrome_service, options=chrome_optiton)

chrome_driver.get(TARGET_URL)
sleep(2)

search_bar = chrome_driver.find_element(By.NAME,"q")
search_bar.send_keys(SEARCH_WORD)
sleep(2)
search_bar.submit()

cnt = 0
results = []
flag = False
while True:
    g_list = chrome_driver.find_elements(By.CLASS_NAME,"g")
    for g in g_list:
        result = {}
        try:
            result["url"] = g.find_element(By.TAG_NAME, "a").get_attribute("href")
            results.append(result)
            cnt +=1
        except:
            break
        # if len(results) >= 0:
        #     flag = True
        #     cnt +=1
        #     break
    if flag:
        break
    sleep(2)
    try:
        chrome_driver.find_element(By.ID,"pnnext").click()
    except:
        break
ws["B1"].value = (f"約{cnt}件")
wb.save(f"xlsx08/search_numbers.xlsx")
wb.close()